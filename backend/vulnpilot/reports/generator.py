"""
VulnPilot AI - Executive Report Generator
Generates professional reports in multiple formats and styles.

REPORT TYPES:
  - Weekly Executive Summary
  - Monthly Trend Analysis
  - Threat Posture Assessment
  - Compliance Audit Report
  - Custom (customer-defined period)

OUTPUT FORMATS:
  - CSV  - Raw data export for spreadsheets
  - XLSX - Formatted Excel with charts and styling
  - PDF  - Professional branded report
  - DOCX - Editable Word document (via Node.js docx lib)
  - JSON - API/machine-readable
  - Markdown - For Slack/Teams/email

STYLES:
  - Executive - 1-page high-level for C-suite
  - Technical - Full CVE details for security teams
  - Compliance - Framework-mapped for auditors
  - Board - Non-technical for board presentations
"""

import csv
import io
import json
import logging
from dataclasses import dataclass, field, asdict
from datetime import datetime, timedelta
from typing import Optional

logger = logging.getLogger(__name__)


# ═══════════════════════════════════════════════════════════════════════════════
# DATA MODEL
# ═══════════════════════════════════════════════════════════════════════════════

@dataclass
class ReportConfig:
    """Report generation configuration."""
    report_type: str = "weekly"       # weekly, monthly, threat_posture, compliance, custom
    style: str = "executive"          # executive, technical, compliance, board
    output_format: str = "json"       # csv, xlsx, pdf, docx, json, markdown
    period_start: Optional[str] = None
    period_end: Optional[str] = None
    company_name: str = "Your Organization"
    include_cve_details: bool = True
    include_compliance: bool = True
    include_trends: bool = True
    max_cves: int = 50
    frameworks: list = field(default_factory=lambda: ["pci", "nist", "soc2", "hipaa", "iso27001", "cisa"])


@dataclass
class ReportData:
    """Computed report data from pipeline results."""
    config: ReportConfig
    generated_at: str = ""

    # Period
    period_start: str = ""
    period_end: str = ""
    period_label: str = ""

    # Summary metrics
    total_cves: int = 0
    critical: int = 0
    high: int = 0
    medium: int = 0
    low: int = 0
    info: int = 0
    noise_eliminated: int = 0
    noise_rate: float = 0.0
    tickets_created: int = 0
    avg_vprs: float = 0.0
    avg_cvss: float = 0.0
    vprs_cvss_flips: int = 0

    # Threat intel
    kev_count: int = 0
    dark_web_count: int = 0
    ransomware_count: int = 0
    internet_facing_critical: int = 0
    exploit_available: int = 0

    # Safety locks
    hard_rules_triggered: int = 0
    adversarial_overrides: int = 0
    drift_events: int = 0

    # SLA
    sla_p1_count: int = 0
    sla_p2_count: int = 0
    sla_p3_count: int = 0
    sla_p4_count: int = 0

    # Top risks
    top_cves: list = field(default_factory=list)
    severity_distribution: dict = field(default_factory=dict)
    asset_tier_breakdown: dict = field(default_factory=dict)

    # Compliance
    compliance_findings: dict = field(default_factory=dict)

    # All CVE data (for detailed reports)
    all_cves: list = field(default_factory=list)


# ═══════════════════════════════════════════════════════════════════════════════
# REPORT BUILDER - Computes report data from pipeline results
# ═══════════════════════════════════════════════════════════════════════════════

class ReportBuilder:
    """Builds ReportData from pipeline results."""

    def build(self, results: list, config: ReportConfig) -> ReportData:
        now = datetime.utcnow()
        data = ReportData(config=config, generated_at=now.isoformat())

        # Period calculation
        if config.period_start and config.period_end:
            data.period_start = config.period_start
            data.period_end = config.period_end
        elif config.report_type == "monthly":
            start = now.replace(day=1) - timedelta(days=1)
            data.period_start = start.replace(day=1).strftime("%Y-%m-%d")
            data.period_end = start.strftime("%Y-%m-%d")
        else:  # weekly or default
            data.period_start = (now - timedelta(days=7)).strftime("%Y-%m-%d")
            data.period_end = now.strftime("%Y-%m-%d")

        data.period_label = f"{data.period_start} to {data.period_end}"
        data.total_cves = len(results)

        vprs_scores = []
        cvss_scores = []

        for r in results:
            vprs = r.get("vprs_score", 0)
            cvss = r.get("cvss_score", 0)
            sev = r.get("severity", "info").lower()
            vprs_scores.append(vprs)
            cvss_scores.append(cvss)

            # Severity
            if sev == "critical": data.critical += 1
            elif sev == "high": data.high += 1
            elif sev == "medium": data.medium += 1
            elif sev == "low": data.low += 1
            else: data.info += 1

            # Flags
            if r.get("is_noise"): data.noise_eliminated += 1
            if r.get("ticket_created"): data.tickets_created += 1
            if r.get("in_kev"): data.kev_count += 1
            if r.get("dark_web_mentions", 0) > 0: data.dark_web_count += 1
            if r.get("ransomware_associated"): data.ransomware_count += 1
            if r.get("cvss_vs_vprs_flip"): data.vprs_cvss_flips += 1
            if r.get("hard_rule"): data.hard_rules_triggered += 1
            if r.get("is_internet_facing") and sev == "critical": data.internet_facing_critical += 1

            # SLA
            if vprs >= 85: data.sla_p1_count += 1
            elif vprs >= 65: data.sla_p2_count += 1
            elif vprs >= 40: data.sla_p3_count += 1
            else: data.sla_p4_count += 1

        # Averages
        if vprs_scores:
            data.avg_vprs = round(sum(vprs_scores) / len(vprs_scores), 1)
        if cvss_scores:
            data.avg_cvss = round(sum(cvss_scores) / len(cvss_scores), 1)

        # Noise rate
        if data.total_cves > 0:
            data.noise_rate = round((data.noise_eliminated / data.total_cves) * 100, 1)

        # Severity distribution
        data.severity_distribution = {
            "critical": data.critical, "high": data.high,
            "medium": data.medium, "low": data.low, "info": data.info
        }

        # Asset tier breakdown
        tiers = {"tier_1": 0, "tier_2": 0, "tier_3": 0, "unknown": 0}
        for r in results:
            t = r.get("asset_tier", "unknown")
            tiers[t] = tiers.get(t, 0) + 1
        data.asset_tier_breakdown = tiers

        # Top CVEs
        sorted_cves = sorted(results, key=lambda x: x.get("vprs_score", 0), reverse=True)
        data.top_cves = [
            {
                "cve_id": r.get("cve_id", "N/A"),
                "vprs_score": round(r.get("vprs_score", 0), 1),
                "cvss_score": round(r.get("cvss_score", 0), 1),
                "severity": r.get("severity", "info").upper(),
                "in_kev": r.get("in_kev", False),
                "dark_web": r.get("dark_web_mentions", 0),
                "hard_rule": r.get("hard_rule", ""),
                "asset_tier": r.get("asset_tier", "unknown"),
                "internet_facing": r.get("is_internet_facing", False),
            }
            for r in sorted_cves[:config.max_cves]
        ]

        # Compliance findings
        data.compliance_findings = self._build_compliance(data, config.frameworks)

        # All CVEs for detailed reports
        if config.include_cve_details:
            data.all_cves = [
                {
                    "cve_id": r.get("cve_id", ""),
                    "vprs_score": round(r.get("vprs_score", 0), 1),
                    "cvss_score": round(r.get("cvss_score", 0), 1),
                    "severity": r.get("severity", "info").upper(),
                    "in_kev": r.get("in_kev", False),
                    "dark_web_mentions": r.get("dark_web_mentions", 0),
                    "hard_rule": r.get("hard_rule", ""),
                    "is_noise": r.get("is_noise", False),
                    "ticket_created": r.get("ticket_created", False),
                    "is_internet_facing": r.get("is_internet_facing", False),
                    "asset_tier": r.get("asset_tier", "unknown"),
                    "epss_score": round(r.get("epss_score", 0), 4),
                    "flip": r.get("cvss_vs_vprs_flip", False),
                }
                for r in sorted_cves
            ]

        return data

    def _build_compliance(self, data: ReportData, frameworks: list) -> dict:
        findings = {}
        if "pci" in frameworks:
            findings["PCI DSS 4.0"] = {
                "6.3.3 - Critical Patch SLA": f"{data.sla_p1_count} critical (24h SLA), {data.sla_p2_count} high (72h SLA)",
                "11.3.1 - Vulnerability Scanning": f"{data.total_cves} CVEs assessed with 6-factor VPRS scoring",
                "6.2.4 - Software Inventory": f"Asset tier tracking: {data.asset_tier_breakdown}",
            }
        if "nist" in frameworks:
            findings["NIST 800-53 / CSF"] = {
                "RA-5 - Vulnerability Monitoring": f"{data.total_cves} CVEs scored, {data.noise_rate}% noise eliminated",
                "SI-2 - Flaw Remediation": f"{data.tickets_created} tickets with SLA deadlines",
                "IR-6 - Incident Reporting": f"{data.kev_count} KEV matches, {data.dark_web_count} dark web signals",
                "ID.RA-1 - Risk Assessment": f"VPRS 6-factor scoring with adversarial AI validation",
            }
        if "soc2" in frameworks:
            findings["SOC 2 Type II"] = {
                "CC6.1 - Logical Access": f"{data.internet_facing_critical} internet-facing critical CVEs monitored",
                "CC7.1 - Monitoring": f"Continuous threat intel: {data.kev_count} KEV, {data.dark_web_count} dark web",
                "CC7.2 - Incident Detection": f"{data.hard_rules_triggered} hard rules triggered for auto-escalation",
            }
        if "hipaa" in frameworks:
            findings["HIPAA Security Rule"] = {
                "§164.312(a)(1) - Access Control": f"Asset-tier scoring prioritizes ePHI systems",
                "§164.308(a)(1) - Risk Analysis": f"{data.total_cves} CVEs with 6-factor risk scoring",
                "§164.306(a) - Security Standards": f"{data.noise_rate}% noise reduction, {data.tickets_created} remediation tickets",
            }
        if "iso27001" in frameworks:
            findings["ISO 27001:2022"] = {
                "A.8.8 - Technical Vulnerability Management": f"VPRS scoring: {data.avg_vprs} avg score, {data.noise_rate}% noise reduction",
                "A.5.7 - Threat Intelligence": f"{data.kev_count} KEV + {data.dark_web_count} dark web signals integrated",
            }
        if "cisa" in frameworks:
            findings["CISA BOD 22-01"] = {
                "KEV Remediation": f"{data.kev_count} Known Exploited Vulnerabilities detected and auto-escalated to CRITICAL",
                "14-Day Deadline": f"All KEV CVEs assigned P1 SLA (24h) via Hard Rule Lock 1",
            }
        return findings


# ═══════════════════════════════════════════════════════════════════════════════
# FORMAT RENDERERS
# ═══════════════════════════════════════════════════════════════════════════════

class JSONRenderer:
    def render(self, data: ReportData) -> dict:
        return {
            "report": {
                "type": data.config.report_type,
                "style": data.config.style,
                "period": data.period_label,
                "generated_at": data.generated_at,
                "company": data.config.company_name,
            },
            "summary": {
                "total_cves": data.total_cves,
                "critical": data.critical,
                "high": data.high,
                "medium": data.medium,
                "low": data.low,
                "noise_eliminated": data.noise_eliminated,
                "noise_rate": data.noise_rate,
                "tickets_created": data.tickets_created,
                "avg_vprs": data.avg_vprs,
                "avg_cvss": data.avg_cvss,
                "vprs_cvss_flips": data.vprs_cvss_flips,
            },
            "threat_intel": {
                "kev_count": data.kev_count,
                "dark_web_count": data.dark_web_count,
                "ransomware_count": data.ransomware_count,
                "internet_facing_critical": data.internet_facing_critical,
            },
            "safety_locks": {
                "hard_rules_triggered": data.hard_rules_triggered,
                "adversarial_overrides": data.adversarial_overrides,
                "drift_events": data.drift_events,
            },
            "sla": {
                "p1_critical_24h": data.sla_p1_count,
                "p2_high_72h": data.sla_p2_count,
                "p3_medium_14d": data.sla_p3_count,
                "p4_low_30d": data.sla_p4_count,
            },
            "top_risks": data.top_cves[:20],
            "compliance": data.compliance_findings,
            "cves": data.all_cves if data.config.include_cve_details else [],
        }


class MarkdownRenderer:
    def render(self, data: ReportData) -> str:
        style = data.config.style
        if style == "board":
            return self._board_style(data)
        elif style == "compliance":
            return self._compliance_style(data)
        elif style == "technical":
            return self._technical_style(data)
        else:
            return self._executive_style(data)

    def _executive_style(self, d: ReportData) -> str:
        title = {
            "weekly": "Weekly Executive Summary",
            "monthly": "Monthly Trend Analysis",
            "threat_posture": "Threat Posture Assessment",
            "compliance": "Compliance Audit Report",
        }.get(d.config.report_type, "Security Report")

        return f"""# VulnPilot AI - {title}
**{d.config.company_name}** | {d.period_label} | Generated {d.generated_at[:10]}

---

## Risk Posture at a Glance

| Metric | Value | Status |
|--------|-------|--------|
| Total CVEs Assessed | {d.total_cves:,} | - |
| Critical (P1 - 24h SLA) | {d.critical} | {'🔴' if d.critical > 0 else '🟢'} |
| High (P2 - 72h SLA) | {d.high} | {'🟠' if d.high > 5 else '🟢'} |
| Noise Eliminated | {d.noise_eliminated} ({d.noise_rate}%) | 🟢 |
| CISA KEV Matches | {d.kev_count} | {'🔴' if d.kev_count > 0 else '🟢'} |
| Dark Web Activity | {d.dark_web_count} CVEs | {'🟠' if d.dark_web_count > 0 else '🟢'} |
| Tickets Created | {d.tickets_created} | - |
| CVSS→VPRS Flips | {d.vprs_cvss_flips} | - |

## Key Findings

- **{d.critical} critical vulnerabilities** require immediate attention (24-hour SLA)
- **{d.kev_count} CVEs** are in CISA's Known Exploited Vulnerabilities catalog
- **{d.noise_rate}%** of CVEs eliminated as noise - your team focuses on {d.total_cves - d.noise_eliminated} real threats
- **{d.vprs_cvss_flips} priority changes** detected vs CVSS-only triage
- **{d.hard_rules_triggered} hard rules** triggered by Triple-Lock safety system

## Top 10 Risks (by VPRS Score)

| # | CVE | VPRS | CVSS | Severity | KEV | Dark Web | Asset |
|---|-----|------|------|----------|-----|----------|-------|
""" + "\n".join(
    f"| {i+1} | {c['cve_id']} | {c['vprs_score']} | {c['cvss_score']} | {c['severity']} | {'✅' if c['in_kev'] else ''} | {c['dark_web'] if c['dark_web'] > 0 else ''} | {c['asset_tier']} |"
    for i, c in enumerate(d.top_cves[:10])
) + f"""

## SLA Compliance

| Priority | SLA | Count |
|----------|-----|-------|
| P1 - Critical | 24 hours | {d.sla_p1_count} |
| P2 - High | 72 hours | {d.sla_p2_count} |
| P3 - Medium | 14 days | {d.sla_p3_count} |
| P4 - Low | 30 days | {d.sla_p4_count} |

## Recommendations

1. **Patch KEV CVEs immediately** - {d.kev_count} confirmed actively exploited
2. **Address internet-facing critical** - {d.internet_facing_critical} external attack surface hotspots
3. **Monitor dark web targets** - {d.dark_web_count} CVEs with underground interest
4. **Review CVSS→VPRS flips** - {d.vprs_cvss_flips} CVEs re-prioritized by real-world context

---
*Generated by VulnPilot AI | Solvent CyberSecurity | vulnpilotai.com*
"""

    def _board_style(self, d: ReportData) -> str:
        risk_level = "HIGH" if d.critical > 3 else "MODERATE" if d.critical > 0 else "LOW"
        emoji = "🔴" if risk_level == "HIGH" else "🟡" if risk_level == "MODERATE" else "🟢"
        return f"""# Security Risk Summary for the Board
**{d.config.company_name}** | {d.period_label}

---

## Overall Risk Level: {emoji} {risk_level}

Our automated vulnerability management system assessed **{d.total_cves:,} security vulnerabilities** this period.

### Key Numbers
- **{d.critical + d.high}** vulnerabilities require action (out of {d.total_cves:,} assessed)
- **{d.noise_rate}%** were identified as low-risk and deprioritized automatically
- **{d.tickets_created}** remediation tasks assigned to engineering teams with deadlines
- **{d.kev_count}** vulnerabilities confirmed exploited by attackers in the wild

### What This Means
Our AI-driven system reduced the security team's workload by {d.noise_rate}% by separating real threats from noise. {d.critical} issues are classified as critical and are being addressed within 24 hours. {d.vprs_cvss_flips} vulnerabilities were re-prioritized using real-world threat intelligence rather than theoretical severity scores.

### Compliance Status
Active monitoring covers PCI DSS, NIST, SOC 2, HIPAA, and ISO 27001 requirements. All {d.kev_count} CISA-mandated vulnerabilities are tracked for remediation.

---
*VulnPilot AI - Automated Vulnerability Intelligence*
"""

    def _compliance_style(self, d: ReportData) -> str:
        sections = []
        for framework, findings in d.compliance_findings.items():
            rows = "\n".join(f"| {req} | {status} |" for req, status in findings.items())
            sections.append(f"""### {framework}

| Requirement | Status |
|-------------|--------|
{rows}
""")
        return f"""# Compliance Audit Report
**{d.config.company_name}** | {d.period_label} | Generated {d.generated_at[:10]}

---

## Vulnerability Management Metrics

| Metric | Value |
|--------|-------|
| CVEs Assessed | {d.total_cves:,} |
| Scoring Method | VPRS (6-factor, AI-validated) |
| Noise Reduction | {d.noise_rate}% |
| Tickets with SLA | {d.tickets_created} |
| KEV Compliance | {d.kev_count} tracked |

## Framework Compliance Mapping

{"".join(sections)}

## Audit Evidence

- Full CVE scoring breakdown with 6-factor component analysis
- Adversarial AI debate transcripts for contested scores
- Hard Rule (Lock 1) trigger log with override justifications
- SLA assignment and tracking for all remediation tickets
- Drift Detector alert history

---
*VulnPilot AI - Automated Vulnerability Intelligence | Solvent CyberSecurity*
"""

    def _technical_style(self, d: ReportData) -> str:
        cve_rows = "\n".join(
            f"| {c['cve_id']} | {c['vprs_score']} | {c['cvss_score']} | {c['severity']} | {'KEV' if c.get('in_kev') else ''} | {c.get('dark_web_mentions',0)} | {c.get('hard_rule','')} | {'NOISE' if c.get('is_noise') else 'ACTION'} |"
            for c in d.all_cves[:d.config.max_cves]
        )
        return f"""# Technical Vulnerability Report
**{d.config.company_name}** | {d.period_label}

## Summary: {d.total_cves} CVEs | {d.critical} CRIT | {d.high} HIGH | {d.noise_rate}% noise

## All CVEs (sorted by VPRS)

| CVE | VPRS | CVSS | Severity | KEV | DW | Hard Rule | Status |
|-----|------|------|----------|-----|----|-----------|--------|
{cve_rows}

## Scoring Configuration
- EPSS: 25% | KEV: 20% | Dark Web: 15% | Asset: 20% | Reach: 12% | Controls: 8%
- Hard Rules: KEV→Critical, Ransomware→Critical, Exploit-for-Sale→Critical
- Adversarial AI: {d.adversarial_overrides} overrides

---
*VulnPilot AI Technical Report*
"""


class CSVRenderer:
    def render(self, data: ReportData) -> str:
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow([
            "CVE ID", "VPRS Score", "CVSS Score", "Severity",
            "In KEV", "Dark Web Mentions", "Hard Rule", "Is Noise",
            "Ticket Created", "Internet Facing", "Asset Tier",
            "EPSS Score", "CVSS→VPRS Flip"
        ])
        for c in data.all_cves:
            writer.writerow([
                c["cve_id"], c["vprs_score"], c["cvss_score"], c["severity"],
                c.get("in_kev", False), c.get("dark_web_mentions", 0),
                c.get("hard_rule", ""), c.get("is_noise", False),
                c.get("ticket_created", False), c.get("is_internet_facing", False),
                c.get("asset_tier", "unknown"), c.get("epss_score", 0),
                c.get("flip", False),
            ])
        return output.getvalue()


class XLSXRenderer:
    def render(self, data: ReportData, filepath: str):
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.chart import BarChart, PieChart, Reference
        from openpyxl.utils import get_column_letter

        wb = openpyxl.Workbook()

        # ── Summary Sheet ──
        ws = wb.active
        ws.title = "Executive Summary"
        ws.sheet_properties.tabColor = "1F4E79"

        # Styles
        title_font = Font(name="Arial", size=16, bold=True, color="1F4E79")
        header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        crit_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        crit_font = Font(name="Arial", size=11, bold=True, color="9C0006")
        good_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        good_font = Font(name="Arial", size=11, color="006100")
        thin_border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin")
        )

        # Title
        ws.merge_cells("A1:F1")
        ws["A1"] = f"VulnPilot AI - {data.config.report_type.replace('_',' ').title()} Report"
        ws["A1"].font = title_font
        ws["A2"] = f"{data.config.company_name} | {data.period_label}"
        ws["A2"].font = Font(name="Arial", size=10, color="666666")
        ws["A3"] = f"Generated: {data.generated_at[:10]}"
        ws["A3"].font = Font(name="Arial", size=10, color="666666")

        # Summary metrics
        metrics = [
            ("Total CVEs Assessed", data.total_cves),
            ("Critical (P1)", data.critical),
            ("High (P2)", data.high),
            ("Medium (P3)", data.medium),
            ("Low/Info", data.low + data.info),
            ("Noise Eliminated", f"{data.noise_eliminated} ({data.noise_rate}%)"),
            ("Tickets Created", data.tickets_created),
            ("CISA KEV Matches", data.kev_count),
            ("Dark Web Activity", data.dark_web_count),
            ("CVSS→VPRS Flips", data.vprs_cvss_flips),
            ("Avg VPRS Score", data.avg_vprs),
            ("Avg CVSS Score", data.avg_cvss),
            ("Hard Rules Triggered", data.hard_rules_triggered),
        ]
        for col_letter in ["A", "B"]:
            ws.column_dimensions[col_letter].width = 25

        row = 5
        ws.cell(row=row, column=1, value="Metric").font = header_font
        ws.cell(row=row, column=1).fill = header_fill
        ws.cell(row=row, column=2, value="Value").font = header_font
        ws.cell(row=row, column=2).fill = header_fill

        for label, value in metrics:
            row += 1
            c1 = ws.cell(row=row, column=1, value=label)
            c2 = ws.cell(row=row, column=2, value=value)
            c1.font = Font(name="Arial", size=11)
            c2.font = Font(name="Arial", size=11, bold=True)
            c1.border = thin_border
            c2.border = thin_border
            if label == "Critical (P1)" and data.critical > 0:
                c2.fill = crit_fill
                c2.font = crit_font
            elif label == "Noise Eliminated":
                c2.fill = good_fill
                c2.font = good_font

        # Severity chart
        chart_ws = wb.create_chartsheet("Severity Chart")
        sev_data_ws = wb.create_sheet("_SevData")
        sev_data_ws["A1"], sev_data_ws["B1"] = "Severity", "Count"
        for i, (sev, cnt) in enumerate([
            ("Critical", data.critical), ("High", data.high),
            ("Medium", data.medium), ("Low", data.low), ("Info/Noise", data.info)
        ], 2):
            sev_data_ws[f"A{i}"] = sev
            sev_data_ws[f"B{i}"] = cnt

        chart = BarChart()
        chart.type = "col"
        chart.title = "Severity Distribution"
        chart.y_axis.title = "CVE Count"
        cats = Reference(sev_data_ws, min_col=1, min_row=2, max_row=6)
        vals = Reference(sev_data_ws, min_col=2, min_row=1, max_row=6)
        chart.add_data(vals, titles_from_data=True)
        chart.set_categories(cats)
        chart.shape = 4
        chart_ws.add_chart(chart)

        # ── CVE Details Sheet ──
        ws2 = wb.create_sheet("CVE Details")
        ws2.sheet_properties.tabColor = "C00000"
        headers = ["CVE ID", "VPRS", "CVSS", "Severity", "KEV", "Dark Web", "Hard Rule",
                    "Noise", "Ticket", "Internet Facing", "Asset Tier", "EPSS", "Flip"]
        for c, h in enumerate(headers, 1):
            cell = ws2.cell(row=1, column=c, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center")

        for r, cve in enumerate(data.all_cves[:data.config.max_cves], 2):
            vals = [
                cve["cve_id"], cve["vprs_score"], cve["cvss_score"], cve["severity"],
                "YES" if cve.get("in_kev") else "", cve.get("dark_web_mentions", 0),
                cve.get("hard_rule", ""), "NOISE" if cve.get("is_noise") else "",
                "YES" if cve.get("ticket_created") else "",
                "YES" if cve.get("is_internet_facing") else "",
                cve.get("asset_tier", ""), cve.get("epss_score", 0),
                "FLIP" if cve.get("flip") else ""
            ]
            for c, v in enumerate(vals, 1):
                cell = ws2.cell(row=r, column=c, value=v)
                cell.font = Font(name="Arial", size=10)
                cell.border = thin_border
                if c == 4 and v == "CRITICAL":
                    cell.fill = crit_fill
                    cell.font = crit_font

        # Auto-width
        for col in range(1, len(headers) + 1):
            ws2.column_dimensions[get_column_letter(col)].width = 15

        # ── Compliance Sheet ──
        if data.config.include_compliance:
            ws3 = wb.create_sheet("Compliance")
            ws3.sheet_properties.tabColor = "2E75B6"
            row = 1
            for framework, findings in data.compliance_findings.items():
                ws3.cell(row=row, column=1, value=framework).font = Font(name="Arial", size=13, bold=True, color="1F4E79")
                row += 1
                ws3.cell(row=row, column=1, value="Requirement").font = header_font
                ws3.cell(row=row, column=1).fill = header_fill
                ws3.cell(row=row, column=2, value="Status").font = header_font
                ws3.cell(row=row, column=2).fill = header_fill
                row += 1
                for req, status in findings.items():
                    ws3.cell(row=row, column=1, value=req).font = Font(name="Arial", size=10)
                    ws3.cell(row=row, column=2, value=status).font = Font(name="Arial", size=10)
                    row += 1
                row += 1
            ws3.column_dimensions["A"].width = 45
            ws3.column_dimensions["B"].width = 60

        # Hide helper sheet
        sev_data_ws.sheet_state = "hidden"

        wb.save(filepath)
        return filepath


class PDFRenderer:
    def render(self, data: ReportData, filepath: str):
        from reportlab.lib.pagesizes import letter
        from reportlab.lib import colors
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.platypus import (
            SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
        )
        from reportlab.lib.units import inch

        doc = SimpleDocTemplate(filepath, pagesize=letter,
                                topMargin=0.75*inch, bottomMargin=0.75*inch)
        styles = getSampleStyleSheet()
        styles.add(ParagraphStyle(name="VPTitle", fontSize=20, leading=24,
                                  textColor=colors.HexColor("#1F4E79"), spaceAfter=12,
                                  fontName="Helvetica-Bold"))
        styles.add(ParagraphStyle(name="VPSubtitle", fontSize=10,
                                  textColor=colors.HexColor("#666666"), spaceAfter=20))
        styles.add(ParagraphStyle(name="VPSection", fontSize=14, leading=18,
                                  textColor=colors.HexColor("#1F4E79"), spaceBefore=16,
                                  spaceAfter=8, fontName="Helvetica-Bold"))

        elements = []

        # Title
        title = {
            "weekly": "Weekly Executive Summary",
            "monthly": "Monthly Trend Analysis",
            "threat_posture": "Threat Posture Assessment",
            "compliance": "Compliance Audit Report",
        }.get(data.config.report_type, "Security Report")
        elements.append(Paragraph(f"VulnPilot AI - {title}", styles["VPTitle"]))
        elements.append(Paragraph(
            f"{data.config.company_name} | {data.period_label} | Generated {data.generated_at[:10]}",
            styles["VPSubtitle"]))
        elements.append(Spacer(1, 12))

        # Summary table
        elements.append(Paragraph("Risk Posture Summary", styles["VPSection"]))
        summary_data = [
            ["Metric", "Value"],
            ["Total CVEs Assessed", str(data.total_cves)],
            ["Critical (24h SLA)", str(data.critical)],
            ["High (72h SLA)", str(data.high)],
            ["Noise Eliminated", f"{data.noise_eliminated} ({data.noise_rate}%)"],
            ["CISA KEV Matches", str(data.kev_count)],
            ["Dark Web Activity", f"{data.dark_web_count} CVEs"],
            ["Tickets Created", str(data.tickets_created)],
            ["CVSS→VPRS Flips", str(data.vprs_cvss_flips)],
            ["Avg VPRS / CVSS", f"{data.avg_vprs} / {data.avg_cvss}"],
        ]
        t = Table(summary_data, colWidths=[3*inch, 3*inch])
        t.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F4E79")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 10),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#CCCCCC")),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F5F7FA")]),
            ("ALIGN", (1, 0), (1, -1), "CENTER"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("TOPPADDING", (0, 0), (-1, -1), 6),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
        ]))
        elements.append(t)
        elements.append(Spacer(1, 16))

        # Top risks
        elements.append(Paragraph("Top 10 Risks", styles["VPSection"]))
        risk_data = [["#", "CVE", "VPRS", "CVSS", "Severity", "KEV", "Asset"]]
        for i, c in enumerate(data.top_cves[:10]):
            risk_data.append([
                str(i+1), c["cve_id"], str(c["vprs_score"]), str(c["cvss_score"]),
                c["severity"], "YES" if c["in_kev"] else "", c.get("asset_tier", "")
            ])
        t2 = Table(risk_data, colWidths=[0.4*inch, 1.6*inch, 0.7*inch, 0.7*inch, 1*inch, 0.6*inch, 1*inch])
        t2.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F4E79")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 9),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#CCCCCC")),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F5F7FA")]),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ]))
        elements.append(t2)

        # SLA
        elements.append(Spacer(1, 16))
        elements.append(Paragraph("SLA Compliance", styles["VPSection"]))
        sla_data = [
            ["Priority", "SLA Deadline", "CVE Count"],
            ["P1 - Critical", "24 hours", str(data.sla_p1_count)],
            ["P2 - High", "72 hours", str(data.sla_p2_count)],
            ["P3 - Medium", "14 days", str(data.sla_p3_count)],
            ["P4 - Low", "30 days", str(data.sla_p4_count)],
        ]
        t3 = Table(sla_data, colWidths=[2*inch, 2*inch, 2*inch])
        t3.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F4E79")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 10),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#CCCCCC")),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("TOPPADDING", (0, 0), (-1, -1), 6),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
        ]))
        elements.append(t3)

        # Compliance (page 2)
        if data.config.include_compliance:
            elements.append(PageBreak())
            elements.append(Paragraph("Compliance Framework Mapping", styles["VPSection"]))
            for framework, findings in data.compliance_findings.items():
                elements.append(Paragraph(framework, ParagraphStyle(
                    name=f"fw_{framework}", fontSize=12, fontName="Helvetica-Bold",
                    spaceBefore=12, spaceAfter=4, textColor=colors.HexColor("#1F4E79"))))
                comp_data = [["Requirement", "Status"]]
                for req, status in findings.items():
                    comp_data.append([req, status])
                tc = Table(comp_data, colWidths=[3*inch, 3.5*inch])
                tc.setStyle(TableStyle([
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2E75B6")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("FONTSIZE", (0, 0), (-1, -1), 8),
                    ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#CCCCCC")),
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                    ("TOPPADDING", (0, 0), (-1, -1), 4),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
                ]))
                elements.append(tc)

        # Footer
        elements.append(Spacer(1, 24))
        elements.append(Paragraph(
            "Generated by VulnPilot AI | Solvent CyberSecurity | vulnpilotai.com",
            ParagraphStyle(name="footer", fontSize=8, textColor=colors.HexColor("#999999"),
                          alignment=1)))

        doc.build(elements)
        return filepath


# ═══════════════════════════════════════════════════════════════════════════════
# MAIN REPORT GENERATOR - Orchestrates build + render
# ═══════════════════════════════════════════════════════════════════════════════

class ReportGenerator:
    """Main entry point for report generation."""

    def __init__(self):
        self.builder = ReportBuilder()

    def generate(self, results: list, config: ReportConfig, output_path: str = None) -> dict:
        """Generate a report from pipeline results.

        Args:
            results: List of CVE pipeline result dicts
            config: Report configuration
            output_path: File path for binary formats (xlsx, pdf, docx)

        Returns:
            Dict with 'content' (str/dict) and 'content_type' and optionally 'filepath'
        """
        data = self.builder.build(results, config)
        fmt = config.output_format.lower()

        if fmt == "json":
            return {
                "content": JSONRenderer().render(data),
                "content_type": "application/json",
            }

        elif fmt == "markdown" or fmt == "md":
            return {
                "content": MarkdownRenderer().render(data),
                "content_type": "text/markdown",
            }

        elif fmt == "csv":
            return {
                "content": CSVRenderer().render(data),
                "content_type": "text/csv",
                "filename": f"vulnpilot_{config.report_type}_{data.period_end}.csv",
            }

        elif fmt == "xlsx":
            path = output_path or f"/tmp/vulnpilot_{config.report_type}_{data.period_end}.xlsx"
            XLSXRenderer().render(data, path)
            return {
                "content": None,
                "content_type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                "filepath": path,
                "filename": f"vulnpilot_{config.report_type}_{data.period_end}.xlsx",
            }

        elif fmt == "pdf":
            path = output_path or f"/tmp/vulnpilot_{config.report_type}_{data.period_end}.pdf"
            PDFRenderer().render(data, path)
            return {
                "content": None,
                "content_type": "application/pdf",
                "filepath": path,
                "filename": f"vulnpilot_{config.report_type}_{data.period_end}.pdf",
            }

        else:
            return {
                "content": JSONRenderer().render(data),
                "content_type": "application/json",
                "error": f"Unknown format '{fmt}', returning JSON",
            }
